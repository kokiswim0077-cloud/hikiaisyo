from __future__ import annotations

import json
import base64
import logging
import mimetypes
import os
import re
import secrets
import shutil
import time
import unicodedata
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from logging.handlers import RotatingFileHandler
from pathlib import Path
from urllib import request

from flask import Flask, Response, jsonify, request as flask_request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
SOURCE_BOOK = Path(os.getenv("INQUIRY_TEMPLATE", str(BASE_DIR / "template.xlsx")))
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", r"C:\Users\koki0\outputs\inquiry_voice_form"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR = Path(os.getenv("LOG_DIR", str(BASE_DIR / "logs")))
LOG_DIR.mkdir(parents=True, exist_ok=True)


def load_local_env() -> None:
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_local_env()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = int(os.getenv("MAX_UPLOAD_MB", "12")) * 1024 * 1024

audit_logger = logging.getLogger("audit")
audit_logger.setLevel(logging.INFO)
audit_handler = RotatingFileHandler(
    LOG_DIR / "audit.log",
    maxBytes=int(os.getenv("AUDIT_LOG_MAX_BYTES", "1048576")),
    backupCount=int(os.getenv("AUDIT_LOG_BACKUPS", "10")),
    encoding="utf-8",
)
audit_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
if not audit_logger.handlers:
    audit_logger.addHandler(audit_handler)

RATE_BUCKETS: dict[tuple[str, str], list[float]] = {}
DOWNLOAD_TOKENS: dict[str, tuple[Path, float]] = {}
LAST_CLEANUP = 0.0
ALLOWED_UPLOAD_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".pdf"}
ALLOWED_UPLOAD_MIME_PREFIXES = ("image/",)
ALLOWED_UPLOAD_MIME_TYPES = {"application/pdf"}
OREC_INTERNAL_CODES = {"kt"}
OREC_INTERNAL_WORDS = ("オーレック", "ｵｰﾚｯｸ", "関東営業所", "谷尾")


def max_text_chars() -> int:
    return int(os.getenv("MAX_TEXT_CHARS", "4000"))


def env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def client_ip() -> str:
    if env_bool("TRUST_PROXY_HEADERS", True):
        cf_ip = flask_request.headers.get("CF-Connecting-IP")
        if cf_ip:
            return cf_ip
        forwarded_for = flask_request.headers.get("X-Forwarded-For", "")
        if forwarded_for:
            return forwarded_for.split(",", 1)[0].strip()
    return flask_request.remote_addr or "unknown"


def audit(event: str, **fields: object) -> None:
    safe_fields = {"event": event, "ip": client_ip(), "path": flask_request.path, **fields}
    audit_logger.info(json.dumps(safe_fields, ensure_ascii=False, default=str))


def json_payload() -> dict[str, object] | None:
    payload = flask_request.get_json(silent=True)
    return payload if isinstance(payload, dict) else None


def allowed_upload_signature(data: bytes, suffix: str) -> bool:
    if suffix == ".pdf":
        return data.startswith(b"%PDF-")
    if suffix in {".jpg", ".jpeg"}:
        return data.startswith(b"\xff\xd8\xff")
    if suffix == ".png":
        return data.startswith(b"\x89PNG\r\n\x1a\n")
    if suffix == ".webp":
        return len(data) >= 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP"
    return False


def cleanup_expired_files() -> None:
    global LAST_CLEANUP
    now = time.time()
    if now - LAST_CLEANUP < 300:
        return
    LAST_CLEANUP = now
    retention_seconds = int(os.getenv("OUTPUT_RETENTION_MINUTES", "120")) * 60
    for token, (path, expires_at) in list(DOWNLOAD_TOKENS.items()):
        if expires_at <= now:
            DOWNLOAD_TOKENS.pop(token, None)
    for path in OUTPUT_DIR.glob("*.xlsx"):
        try:
            if now - path.stat().st_mtime > retention_seconds:
                path.unlink()
                audit_logger.info(json.dumps({"event": "cleanup_deleted", "path": str(path)}, ensure_ascii=False))
        except OSError as exc:
            audit_logger.warning(json.dumps({"event": "cleanup_failed", "path": str(path), "error": str(exc)}, ensure_ascii=False))


def rate_limit_key() -> str:
    if flask_request.path == "/api/parse-image":
        return "image"
    if flask_request.path == "/api/parse":
        return "parse"
    if flask_request.path == "/api/save":
        return "save"
    if flask_request.path.startswith("/download/"):
        return "download"
    return "default"


def rate_limit_config(bucket: str) -> tuple[int, int]:
    configs = {
        "image": (int(os.getenv("RATE_IMAGE_COUNT", "12")), int(os.getenv("RATE_IMAGE_WINDOW", "600"))),
        "parse": (int(os.getenv("RATE_PARSE_COUNT", "60")), int(os.getenv("RATE_PARSE_WINDOW", "60"))),
        "save": (int(os.getenv("RATE_SAVE_COUNT", "30")), int(os.getenv("RATE_SAVE_WINDOW", "60"))),
        "download": (int(os.getenv("RATE_DOWNLOAD_COUNT", "60")), int(os.getenv("RATE_DOWNLOAD_WINDOW", "60"))),
        "default": (int(os.getenv("RATE_DEFAULT_COUNT", "180")), int(os.getenv("RATE_DEFAULT_WINDOW", "60"))),
    }
    return configs.get(bucket, configs["default"])


def is_rate_limited() -> bool:
    bucket = rate_limit_key()
    limit, window = rate_limit_config(bucket)
    key = (client_ip(), bucket)
    now = time.time()
    timestamps = [ts for ts in RATE_BUCKETS.get(key, []) if now - ts < window]
    if len(timestamps) >= limit:
        RATE_BUCKETS[key] = timestamps
        audit("rate_limited", bucket=bucket, limit=limit, window=window)
        return True
    timestamps.append(now)
    RATE_BUCKETS[key] = timestamps
    return False


def request_is_https() -> bool:
    if flask_request.is_secure:
        return True
    return flask_request.headers.get("X-Forwarded-Proto", "").lower() == "https"


def app_password() -> str:
    return os.getenv("APP_PASSWORD", "")


def is_authenticated() -> bool:
    password = app_password()
    if not password:
        return True
    auth = flask_request.authorization
    return bool(auth and auth.username == "user" and secrets.compare_digest(auth.password or "", password))


@app.before_request
def require_app_password():
    cleanup_expired_files()
    if env_bool("PUBLIC_MODE", False) and not app_password():
        return jsonify({"error": "APP_PASSWORD is required in PUBLIC_MODE"}), 503
    if env_bool("REQUIRE_HTTPS", False) and not request_is_https():
        audit("blocked_non_https")
        return jsonify({"error": "HTTPS required"}), 403
    if is_rate_limited():
        return jsonify({"error": "Too many requests. Please wait and retry."}), 429
    if flask_request.path == "/api/status" and not env_bool("PUBLIC_MODE", False):
        return None
    if is_authenticated():
        return None
    audit("auth_failed")
    return Response(
        "Authentication required",
        401,
        {"WWW-Authenticate": 'Basic realm="Inquiry Voice Form"'},
    )


def today() -> datetime:
    return datetime.now()


def add_business_days(start: datetime, days: int) -> datetime:
    current = start
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:
            added += 1
    return current


def date_from_iso(value: str) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


def normalize(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.lower()
    return re.sub(r"[\s　・･()（）株式会社有限会社㈱㈲\-_/.,、。]+", "", text)


def load_master_data() -> dict[str, list[dict[str, object]]]:
    wb = load_workbook(SOURCE_BOOK, read_only=True, data_only=True)

    customers: list[dict[str, object]] = []
    for code, name, kana, rate, *_ in wb["請求先"].iter_rows(min_row=2, values_only=True):
        if code and name:
            customers.append(
                {
                    "code": str(code),
                    "name": str(name).strip(),
                    "kana": "" if kana is None else str(kana).strip(),
                    "rate": rate,
                }
            )

    deliveries: list[dict[str, object]] = []
    for code, kana, name, postal, address, tel, *_ in wb["出荷先"].iter_rows(min_row=2, values_only=True):
        if code and name:
            deliveries.append(
                {
                    "code": str(code),
                    "name": str(name).strip(),
                    "kana": "" if kana is None else str(kana).strip(),
                    "postal": "" if postal is None else str(postal).strip(),
                    "address": "" if address is None else str(address).strip(),
                    "tel": "" if tel is None else str(tel).strip(),
                }
            )

    products: list[dict[str, object]] = []
    for code, name, price, *_ in wb["商品コード"].iter_rows(min_row=10, values_only=True):
        if code and name:
            products.append({"code": str(code), "name": str(name).strip(), "price": price})

    return {"customers": customers, "deliveries": deliveries, "products": products}


MASTER = load_master_data()


def best_match(query: str, rows: list[dict[str, object]], keys: tuple[str, ...]) -> dict[str, object] | None:
    candidates = ranked_matches(query, rows, keys, limit=1)
    return candidates[0] if candidates else None


def ranked_matches(
    query: str,
    rows: list[dict[str, object]],
    keys: tuple[str, ...],
    limit: int = 8,
) -> list[dict[str, object]]:
    q = normalize(query)
    if not q:
        return []

    scored: list[dict[str, object]] = []
    for row in rows:
        haystacks = [normalize(row.get(key, "")) for key in keys]
        score = 0.0
        for hay in haystacks:
            if not hay:
                continue
            if q == hay:
                score = max(score, 1.0)
            elif q in hay or hay in q:
                score = max(score, 0.92)
            else:
                score = max(score, SequenceMatcher(None, q, hay).ratio())
        if score >= 0.45:
            result = dict(row)
            result["score"] = round(score, 3)
            scored.append(result)

    scored.sort(key=lambda item: item["score"], reverse=True)
    return scored[:limit]


def needs_confirmation(candidates: list[dict[str, object]]) -> bool:
    if len(candidates) < 2:
        return False
    top = float(candidates[0].get("score", 0))
    second = float(candidates[1].get("score", 0))
    if top - second <= 0.12:
        return True
    top_name = normalize(candidates[0].get("name", ""))
    for candidate in candidates[1:4]:
        name = normalize(candidate.get("name", ""))
        if top_name and name and (top_name in name or name in top_name):
            return True
    return False


def extract_after_keyword(text: str, keywords: list[str], stops: list[str]) -> str:
    fragment = extract_labeled_fragment(text, keywords, stops)
    return "" if fragment is None else fragment


def extract_labeled_fragment(text: str, keywords: list[str], stops: list[str]) -> str | None:
    for keyword in keywords:
        m = re.search(keyword, text)
        if not m:
            continue
        tail = text[m.end() :]
        stop_positions = [tail.find(stop) for stop in stops if tail.find(stop) >= 0]
        if stop_positions:
            tail = tail[: min(stop_positions)]
        return tail.strip(" 、。,.")
    return None


def clean_query_fragment(value: str) -> str:
    value = re.sub(r"^[、。\s,]+", "", value)
    parts = [part.strip() for part in re.split(r"[、。\n,]", value) if part.strip()]
    return parts[0] if parts else value.strip()


def is_orec_internal(value: object) -> bool:
    text = str(value or "")
    return any(word in text for word in OREC_INTERNAL_WORDS)


def infer_product_query(text: str) -> str:
    explicit = extract_after_keyword(
        text,
        ["製品", "商品", "品番", "型式", "機種"],
        ["数量", "倉庫", "値引", "受注日", "日中日", "注文日", "出荷希望日", "出荷日"],
    )
    explicit = clean_query_fragment(explicit)
    if explicit:
        return explicit

    tokens = [token.strip() for token in re.split(r"[、。\s,\n]+", text) if token.strip()]
    best_token = ""
    best_score = 0.0
    for token in tokens:
        normalized = unicodedata.normalize("NFKC", token).upper()
        if not re.search(r"[A-Z]", normalized) or not re.search(r"\d", normalized):
            continue
        candidates = ranked_matches(normalized, MASTER["products"], ("name", "code"), limit=1)
        if candidates and float(candidates[0].get("score", 0)) > best_score:
            best_token = normalized
            best_score = float(candidates[0].get("score", 0))
    return best_token if best_score >= 0.65 else ""


def model_tokens_from_text(text: str) -> list[str]:
    normalized = unicodedata.normalize("NFKC", text or "").upper()
    tokens = re.findall(r"\b[A-Z]{1,5}\d{2,4}[A-Z]{0,3}(?:/[A-Z0-9]+)?\b", normalized)
    tokens.extend(re.findall(r"\b[A-Z]{1,5}\d{2,4}(?:/[A-Z0-9]+)?\b", normalized))
    # Prefer more specific tokens first: RCHR800A before RCHR800, RM983FX before RM983.
    return sorted(dict.fromkeys(tokens), key=len, reverse=True)


def parse_date(text: str, label_patterns: list[str]) -> str:
    fragment = extract_labeled_fragment(
        text,
        label_patterns,
        ["得意先", "特異先", "納入先", "倉庫", "値引", "製品", "商品", "受注日", "日中日", "注文日", "出荷希望日", "出荷日"],
    )
    if fragment is None:
        return ""
    base = today()
    target = fragment
    if "明後日" in target or "あさって" in target:
        return (base + timedelta(days=2)).strftime("%Y-%m-%d")
    if "明日" in target:
        return (base + timedelta(days=1)).strftime("%Y-%m-%d")
    if "今日" in target or "本日" in target:
        return base.strftime("%Y-%m-%d")

    m = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", target)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.search(r"(\d{1,2})月(\d{1,2})日", target)
    if m:
        return f"{base.year:04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return ""


def resolve_default_dates(parsed: dict[str, object]) -> dict[str, object]:
    result = dict(parsed)
    if not result.get("order_date"):
        result["order_date"] = today().strftime("%Y-%m-%d")

    production_date = str(result.get("production_date") or "")
    ship_date = str(result.get("ship_date") or "")
    ship_text = str(result.get("ship_text") or "")
    production_dt = date_from_iso(production_date)
    if not ship_date and not ship_text and production_dt:
        result["ship_date"] = add_business_days(production_dt, 2).strftime("%Y-%m-%d")
    return result


def normalize_warehouse(value: object) -> str:
    code = str(value or "").strip()
    if re.fullmatch(r"0?\d{2,3}", code):
        code = code.zfill(3)
    if code == "033":
        return "031"
    return code if code in {"011", "031"} else ""


def sanitize_parsed(parsed: dict[str, object]) -> dict[str, object]:
    result = dict(parsed)
    result["warehouse"] = normalize_warehouse(result.get("warehouse"))
    if result.get("discount_method") not in {"外掛", "内掛", ""}:
        text = str(result.get("discount_method") or "")
        if re.search(r"外掛|外がけ|外掛け|外書き", text):
            result["discount_method"] = "外掛"
        elif re.search(r"内掛|内がけ|内掛け|内書き", text):
            result["discount_method"] = "内掛"
        else:
            result["discount_method"] = ""
    rate = str(result.get("discount_rate") or "")
    m = re.search(r"\d+(?:\.\d+)?", rate)
    result["discount_rate"] = m.group(0) if m else ""
    try:
        result["quantity"] = int(result.get("quantity") or 1)
    except (TypeError, ValueError):
        result["quantity"] = 1
    return result


def local_parse(text: str) -> dict[str, object]:
    compact = unicodedata.normalize("NFKC", text)
    stops = ["受注日", "日中日", "注文日", "出荷希望日", "出荷日", "納入先", "倉庫", "値引", "製品", "商品", "品番", "型式", "機種", "数量"]

    customer_query = clean_query_fragment(extract_after_keyword(compact, ["得意先", "特異先", "取引先", "お客"], stops))
    delivery_query = clean_query_fragment(extract_after_keyword(compact, ["納入先", "出荷先"], stops))
    if delivery_query in {"得意先", "特異先"}:
        delivery_query = ""
    product_query = infer_product_query(compact)

    warehouse = ""
    m = re.search(r"倉庫\s*(0?\d{2,3})", compact)
    if m:
        warehouse = normalize_warehouse(m.group(1))
    if not warehouse:
        if "011" in compact or "福岡倉庫" in compact or "本社" in compact:
            warehouse = "011"
        elif "031" in compact or "033" in compact or any(term in compact for term in ["オーレック関物", "関東物流", "関物", "関東物"]):
            warehouse = "031"

    discount_method = ""
    if re.search(r"外掛|外がけ|外掛け|外書き", compact):
        discount_method = "外掛"
    elif re.search(r"内掛|内がけ|内掛け|内書き", compact):
        discount_method = "内掛"

    discount_rate = ""
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", compact)
    if not m:
        m = re.search(r"値引(?:き)?(?:率)?\s*(\d+(?:\.\d+)?)\s*(?:パーセント|パー)?", compact)
    if m:
        discount_rate = m.group(1)

    discount_name = ""
    for name in ["実演機対応値引", "実演機値引", "展示機値引", "不需要期値引", "早期確注値引", "台数値引", "推進値引", "展示会値引"]:
        if name in compact or name.replace("値引", "値引き") in compact:
            discount_name = name
            break

    quantity = 1
    m = re.search(r"(?:数量\s*)?(\d+)\s*(?:台|個|本|枚)", compact)
    if m:
        quantity = int(m.group(1))

    return {
        "customer_query": customer_query,
        "delivery_query": delivery_query,
        "order_date": parse_date(compact, ["受注日", "日中日", "注文日"]),
        "production_date": parse_date(compact, ["生産日", "生産", "製造日"]),
        "ship_date": parse_date(compact, ["出荷希望日", "出荷日", "希望日"]),
        "warehouse": warehouse,
        "discount_name": discount_name,
        "discount_method": discount_method,
        "discount_rate": discount_rate,
        "product_query": product_query,
        "quantity": quantity,
    }


def request_gemini_json(endpoint: str, body: dict[str, object], api_key: str, timeout: int) -> dict[str, object] | None:
    encoded_body = json.dumps(body).encode("utf-8")
    for attempt in range(2):
        req = request.Request(
            endpoint,
            data=encoded_body,
            headers={"Content-Type": "application/json", "x-goog-api-key": api_key},
            method="POST",
        )
        try:
            with request.urlopen(req, timeout=timeout) as res:
                data = json.loads(res.read().decode("utf-8"))
            text_part = data["candidates"][0]["content"]["parts"][0]["text"]
            return json.loads(text_part)
        except Exception:
            if attempt == 1:
                return None
            time.sleep(0.6)
    return None


def parse_with_gemini(text: str, api_key: str) -> dict[str, object] | None:
    model = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    prompt = f"""
日本語の音声認識テキストから、引合書入力用の項目をJSONだけで抽出してください。
不明な項目は空文字にしてください。日付は今日={today().strftime('%Y-%m-%d')}を基準にYYYY-MM-DDへ変換してください。
キー:
customer_query, delivery_query, order_date, production_date, ship_date, warehouse, discount_name, discount_method, discount_rate, product_query, quantity
discount_methodは外掛または内掛。音声誤認識の「内書き」は内掛として扱ってください。warehouseは011または031。033と読めた場合は031として扱ってください。discount_rateは3%なら3。

テキスト:
{text}
""".strip()
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"responseMimeType": "application/json"},
    }
    return request_gemini_json(endpoint, body, api_key, timeout=20)


def parse_image_with_gemini(image_bytes: bytes, filename: str, api_key: str) -> dict[str, object] | None:
    model = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    mime_type = mimetypes.guess_type(filename)[0] or "image/jpeg"
    prompt = f"""
注文書・発注書の画像を読み取り、引合書入力用の項目をJSONだけで抽出してください。
読み取れない項目は空文字にしてください。日付は今日={today().strftime('%Y-%m-%d')}を基準にYYYY-MM-DDへ正規化してください。
品名・型式・品番から製品候補を読み取ってください。得意先と納入先が同じとは限りません。

キー:
document_type, issuer_query, visible_text, customer_query, customer_code, delivery_query, delivery_code, delivery_office_name, delivery_office_code, order_date, production_date, production_text, ship_date, ship_text, warehouse, discount_name, discount_method, discount_rate, product_query, product_code, quantity, order_no

warehouseは011または031。不明なら空文字。033と読めた場合は031として扱ってください。
discount_methodは外掛または内掛。不明なら空文字。
discount_nameは台数値引、不需要期値引、実演機対応値引などの値引名。不明なら空文字。
discount_rateは3%なら3。
quantityは数値。
production_date/ship_dateは日まで読める場合だけYYYY-MM-DD。月だけの場合は production_text/ship_text に「8月」「2026年9月」「2026年10月以降」のように入れる。
issuer_queryは帳票右上や発行元に書かれた会社名。visible_textは判断に重要な見出し・会社名・備考欄だけを短くまとめる。

クボタの「注文書（出荷指示書）」形式の場合:
- 右上に株式会社クボタ、得意先名に関東甲信クボタがある場合は、customer_queryは「関東甲信クボタ（千葉県）」として扱う。
- 「納所名」「納所コード」「届先」欄を重視する。例: 納所名が「大網営業所」なら delivery_office_name は「大網営業所」。
- 下部の備考欄に「011倉庫」「6/29(月)出荷」のような手書き/赤字がある場合、warehouseとship_dateはそこから読む。
- 商品は「形式名」「型式名」「商品名」欄を重視する。例: RM953X/K。

オーレックの「引き合い内容連絡書」形式の場合:
- 得意先NO/納入先NOを優先する。
- 出荷希望日が手書き備考にある場合はそこを優先する。

一般の「注文書」形式の場合:
- 右上に書かれた注文元会社を customer_query として読む。例: 有限会社 木嶋商店。
- 備考欄に「生産日 8月」「出荷予定日 2026年9月」のように月だけがある場合、production_text/ship_text に入れる。

「受注・発注カード」形式の場合:
- 右上の会社名を得意先として読む。例: 株式会社エルタ。
- 表内の得意先名や直送先名は納入先・直送先の候補であり、得意先として優先しない。
- 備考欄に「出荷予定日 2026年 10月 以降」のように月だけがある場合、ship_text に入れる。
""".strip()
    body = {
        "contents": [
            {
                "parts": [
                    {"text": prompt},
                    {
                        "inlineData": {
                            "mimeType": mime_type,
                            "data": base64.b64encode(image_bytes).decode("ascii"),
                        }
                    },
                ]
            }
        ],
        "generationConfig": {"responseMimeType": "application/json"},
    }
    return request_gemini_json(endpoint, body, api_key, timeout=25)


def row_by_code(rows: list[dict[str, object]], code: object) -> dict[str, object] | None:
    wanted = str(code or "").strip()
    if not wanted:
        return None
    for row in rows:
        if str(row.get("code", "")).strip() == wanted:
            return {**row, "score": 1.0}
    return None


def apply_shipping_source_warehouse(parsed: dict[str, object]) -> dict[str, object]:
    result = dict(parsed)
    blob = " ".join(str(value or "") for value in result.values())

    m = re.search(r"\b(011|031|033)\s*倉庫", blob)
    if m:
        result["warehouse"] = normalize_warehouse(m.group(1))

    if any(term in blob for term in ["オーレック関物", "関東物流", "関物", "関東物"]):
        result["warehouse"] = "031"
    elif any(term in blob for term in ["福岡倉庫", "オーレック本社", "本社 福岡", "本社（福岡）", "本社(福岡)"]):
        result["warehouse"] = "011"

    return result


def kubota_order_overrides(parsed: dict[str, object]) -> dict[str, object]:
    result = dict(parsed)
    blob = " ".join(str(value or "") for value in result.values())

    is_kubota_shipping_order = (
        "関東甲信クボタ" in blob
        or "株式会社クボタ" in blob
        or "クボタ" in blob and ("営業所" in blob or "納所" in blob or "注文書" in blob)
    )
    if not is_kubota_shipping_order:
        return result

    # Business rule from the user: Kubota order-form format uses customer 61110.
    result["customer_code"] = "61110"
    result["customer_query"] = "関東甲信クボタ（千葉県）"
    result["customer_force_confirmed"] = True
    result["order_date"] = today().strftime("%Y-%m-%d")

    office_name = str(result.get("delivery_office_name") or result.get("delivery_query") or "")
    office_code = str(result.get("delivery_office_code") or "").strip().zfill(3)
    kubota_office_codes = {
        "038": ("61110004", "関東甲信クボタ 市原営業所"),
        "040": ("61110005", "関東甲信クボタ 大網営業所"),
    }
    if "大網" in office_name or "大網" in blob:
        office_code = "040"
    elif "市原" in office_name or "市原" in blob:
        office_code = "038"

    if office_code in kubota_office_codes:
        delivery_code, delivery_query = kubota_office_codes[office_code]
        result["delivery_code"] = delivery_code
        result["delivery_query"] = delivery_query
        result["delivery_force_confirmed"] = True

    if office_code == "034" or "君津" in office_name or "君津" in blob:
        result["delivery_code"] = "61110011"
        result["delivery_query"] = "関東甲信クボタ 君津営業所"
        result["delivery_force_confirmed"] = True

    product_query = str(result.get("product_query") or "")
    if not product_query:
        # Common Kubota form pattern: model names are alphanumeric with a slash.
        m = re.search(r"\b[A-Z]{1,4}\d{2,4}[A-Z]?(?:/[A-Z0-9]+)?\b", blob.upper())
        if m:
            result["product_query"] = m.group(0)

    return result


def known_form_overrides(parsed: dict[str, object]) -> dict[str, object]:
    result = dict(parsed)
    blob = " ".join(str(value or "") for value in result.values())

    if "大竹産業" in blob:
        result["customer_code"] = "61323"
        result["customer_query"] = "大竹産業"
        result["customer_force_confirmed"] = True
        if not result.get("delivery_code") or str(result.get("delivery_code") or "").lower() in OREC_INTERNAL_CODES or is_orec_internal(result.get("delivery_query")):
            result["delivery_code"] = "61323"
            result["delivery_query"] = "大竹産業"
            result["delivery_force_confirmed"] = True

    if "木嶋" in blob or "木島" in blob:
        result["customer_code"] = "61310"
        result["customer_query"] = "木嶋商店"
        result["customer_force_confirmed"] = True
        if not result.get("delivery_code"):
            result["delivery_code"] = "61310"
            result["delivery_query"] = "木嶋商店"
            result["delivery_force_confirmed"] = True

    if "エルタ" in blob or "ｴﾙﾀ" in blob:
        result["customer_code"] = "65137"
        result["customer_query"] = "エルタ"
        result["customer_force_confirmed"] = True
        if not result.get("delivery_code"):
            result["delivery_code"] = "65137"
            result["delivery_query"] = "エルタ"
            result["delivery_force_confirmed"] = True

    return result


def apply_delivery_source_overrides(parsed: dict[str, object]) -> dict[str, object]:
    result = dict(parsed)
    delivery_code = str(result.get("delivery_code") or "").strip().lower()
    delivery_query = str(result.get("delivery_query") or "")
    customer_code = str(result.get("customer_code") or "").strip()
    customer_query = str(result.get("customer_query") or "").strip()

    if delivery_code in OREC_INTERNAL_CODES or is_orec_internal(delivery_query):
        result["delivery_code"] = customer_code
        result["delivery_query"] = customer_query
        if customer_code:
            result["delivery_force_confirmed"] = True

    return result


def resolve_fields(parsed: dict[str, object]) -> dict[str, object]:
    parsed = sanitize_parsed(parsed)
    parsed = kubota_order_overrides(parsed)
    parsed = known_form_overrides(parsed)
    parsed = apply_delivery_source_overrides(parsed)
    parsed = apply_shipping_source_warehouse(parsed)
    parsed = resolve_default_dates(parsed)
    exact_customer = row_by_code(MASTER["customers"], parsed.get("customer_code"))
    if exact_customer:
        customer_candidates = [exact_customer]
    else:
        customer_candidates = ranked_matches(
            str(parsed.get("customer_query", "")),
            MASTER["customers"],
            ("name", "kana", "code"),
        )
    customer = customer_candidates[0] if customer_candidates else None
    exact_delivery = row_by_code(MASTER["deliveries"], parsed.get("delivery_code"))
    delivery_query = str(parsed.get("delivery_query", ""))
    if exact_delivery:
        delivery_candidates = [exact_delivery]
    elif delivery_query:
        delivery_candidates = ranked_matches(delivery_query, MASTER["deliveries"], ("name", "kana", "code", "address"))
    elif customer:
        delivery_candidates = [
            {**row, "score": 1.0}
            for row in MASTER["deliveries"]
            if str(row.get("code")) == str(customer.get("code"))
        ]
    else:
        delivery_candidates = []
    delivery = delivery_candidates[0] if delivery_candidates else None
    exact_product = row_by_code(MASTER["products"], parsed.get("product_code"))
    if exact_product:
        product_candidates = [exact_product]
    else:
        product_search_text = " ".join(
            str(parsed.get(key, "") or "")
            for key in ["product_code", "product_query", "visible_text"]
        )
        token_candidates = []
        for token in model_tokens_from_text(product_search_text):
            token_candidates = ranked_matches(token, MASTER["products"], ("name", "code"))
            if token_candidates and float(token_candidates[0].get("score", 0)) >= 0.9:
                break
        product_candidates = token_candidates or ranked_matches(str(parsed.get("product_query", "")), MASTER["products"], ("name", "code"))
    product = product_candidates[0] if product_candidates else None

    return {
        **parsed,
        "customer": customer,
        "customer_candidates": customer_candidates,
        "customer_name": customer.get("name", "") if customer else str(parsed.get("customer_query", "") or ""),
        "customer_code": customer.get("code", "") if customer else str(parsed.get("customer_code", "") or ""),
        "customer_needs_confirmation": False if parsed.get("customer_force_confirmed") else needs_confirmation(customer_candidates),
        "delivery": delivery,
        "delivery_candidates": delivery_candidates,
        "delivery_name": delivery.get("name", "") if delivery else str(parsed.get("delivery_query", "") or ""),
        "delivery_code": delivery.get("code", "") if delivery else str(parsed.get("delivery_code", "") or ""),
        "delivery_needs_confirmation": False if parsed.get("delivery_force_confirmed") else needs_confirmation(delivery_candidates),
        "product": product,
        "product_candidates": product_candidates,
        "product_name": product.get("name", "") if product else str(parsed.get("product_query", "") or ""),
        "product_code": product.get("code", "") if product else str(parsed.get("product_code", "") or ""),
        "product_needs_confirmation": needs_confirmation(product_candidates),
        "warehouse": normalize_warehouse(parsed.get("warehouse")) or "011",
        "discount_method": parsed.get("discount_method") or "外掛",
        "discount_rate": parsed.get("discount_rate") or "0",
        "quantity": parsed.get("quantity") or 1,
    }


def safe_filename_part(value: object, fallback: str) -> str:
    text = str(value or "").strip() or fallback
    text = re.sub(r'[\\/:*?"<>|]+', "", text)
    text = re.sub(r"\s+", "", text)
    text = text.strip(" ._")
    return text[:40] or fallback


def output_excel_path(fields: dict[str, object]) -> Path:
    delivery = fields.get("delivery") or {}
    product = fields.get("product") or {}
    delivery_name = safe_filename_part(
        delivery.get("name") or fields.get("delivery_name") or fields.get("delivery_query"),
        "納入先未指定",
    )
    product_name = safe_filename_part(
        product.get("name") or fields.get("product_name") or fields.get("product_query"),
        "機種未指定",
    )
    base_name = f"{datetime.now().strftime('%y%m%d')}_{delivery_name}_{product_name}_引合書"
    path = OUTPUT_DIR / f"{base_name}.xlsx"
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = OUTPUT_DIR / f"{base_name}_{index}.xlsx"
        if not candidate.exists():
            return candidate
    return OUTPUT_DIR / f"{base_name}_{secrets.token_hex(4)}.xlsx"


def numeric_value(value: object) -> int:
    text = str(value or "").replace(",", "").strip()
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def coerce_parsed_dict(value: object) -> dict[str, object]:
    if isinstance(value, dict):
        return value
    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                return item
    return {}


def quote_output_path(fields: dict[str, object], suffix: str) -> Path:
    customer = fields.get("customer") or {}
    product = fields.get("product") or {}
    customer_name = safe_filename_part(
        customer.get("name") or fields.get("customer_name") or fields.get("customer_query"),
        "得意先未指定",
    )
    product_name = safe_filename_part(
        product.get("name") or fields.get("product_name") or fields.get("product_query"),
        "機種未指定",
    )
    base_name = f"{datetime.now().strftime('%y%m%d')}_{customer_name}_{product_name}_{suffix}"
    path = OUTPUT_DIR / f"{base_name}.xlsx"
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = OUTPUT_DIR / f"{base_name}_{index}.xlsx"
        if not candidate.exists():
            return candidate
    return OUTPUT_DIR / f"{base_name}_{secrets.token_hex(4)}.xlsx"


def quote_local_parse(text: str) -> dict[str, object]:
    compact = unicodedata.normalize("NFKC", text)
    stops = ["担当", "商品", "製品", "品番", "型式", "機種", "数量", "台数", "在庫", "生産", "出荷", "備考", "値引", "小売", "仕切"]
    customer_query = clean_query_fragment(extract_after_keyword(compact, ["得意先", "お客様", "客先", "宛先"], stops))
    product_query = infer_product_query(compact)

    quantity = 1
    m = re.search(r"(?:数量|台数)?\s*(\d+)\s*(?:台|個|本|枚)", compact)
    if m:
        quantity = int(m.group(1))

    discount_rate = ""
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", compact)
    if not m:
        m = re.search(r"値引(?:き)?(?:率)?\s*(\d+(?:\.\d+)?)\s*(?:パーセント|パー)?", compact)
    if m:
        discount_rate = m.group(1)

    stock_status = ""
    for status in ["在庫限り", "生産待ち", "在庫有り", "有り", "あり", "無し", "なし"]:
        if status in compact:
            stock_status = "有り" if status == "あり" else ("無し" if status == "なし" else status)
            break

    return {
        "customer_query": customer_query,
        "staff_name": clean_query_fragment(extract_after_keyword(compact, ["担当", "担当者"], stops)),
        "quote_date": parse_date(compact, ["見積日", "作成日", "依頼日"]),
        "product_query": product_query,
        "quantity": quantity,
        "retail_price": numeric_value(extract_after_keyword(compact, ["小売単価", "小売"], stops)),
        "wholesale_price": numeric_value(extract_after_keyword(compact, ["仕切単価", "仕切"], stops)),
        "discounted_price": numeric_value(extract_after_keyword(compact, ["値引後単価", "値引後"], stops)),
        "discount_name": clean_query_fragment(extract_after_keyword(compact, ["値引名"], stops)),
        "discount_rate": discount_rate,
        "stock_status": stock_status,
        "production_date": parse_date(compact, ["生産日", "生産"]),
        "ship_date": parse_date(compact, ["出荷予定日", "出荷日", "出荷"]),
        "note": clean_query_fragment(extract_after_keyword(compact, ["備考", "メモ"], [])),
    }


def clean_quote_free_text(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    boilerplate_patterns = [
        r"(?:尚[、,]?\s*)?生産予定日は諸事情により変動する場合[がも]ございますのでご了承(?:ください|願います)。?",
        r"見積り?時点での商品確保はできかねます。?",
        r"商品確保はできかねます。?",
        r"お手数ですが在庫の有無は都度[ご]?確認いただくようお願い申し上げます。?",
        r"出荷日は生産日の翌々日対応です。?",
        r"\d{1,2}/\d{1,2}(?:\([月火水木金土日]\))?\s*早めに返信(?:願います|お願いします)?。?",
        r"早めに返信(?:願います|お願いします)?。?",
        r"毎度格別のお引き立てを賜りお礼申し上げます。?",
        r"下記の通りお見積り申し上げます。?",
    ]
    for pattern in boilerplate_patterns:
        text = re.sub(pattern, "", text)
    text = re.sub(r"[\s　]*\n[\s　]*", "\n", text)
    text = re.sub(r"[ 　]{2,}", " ", text)
    return text.strip(" \n、。,※")


def clean_quote_auxiliary_fields(result: dict[str, object]) -> dict[str, object]:
    cleaned = dict(result)
    for key in ["note", "production_text", "ship_text"]:
        cleaned[key] = clean_quote_free_text(cleaned.get(key))
    if "諸事情" in str(cleaned.get("production_text") or ""):
        cleaned["production_text"] = ""
    if "翌々日対応" in str(cleaned.get("ship_text") or ""):
        cleaned["ship_text"] = ""
    stock_status = str(cleaned.get("stock_status") or "")
    if "在庫限り" in stock_status:
        cleaned["stock_status"] = "在庫限り"
    elif "生産待ち" in stock_status:
        cleaned["stock_status"] = "生産待ち"
    elif "有り" in stock_status or "あり" in stock_status:
        cleaned["stock_status"] = "有り"
    elif "無し" in stock_status or "なし" in stock_status:
        cleaned["stock_status"] = "無し"
    elif "商品確保" in stock_status:
        cleaned["stock_status"] = ""
    if "在庫限り" in str(cleaned.get("note") or "") and cleaned.get("stock_status") in {"", "有り"}:
        cleaned["stock_status"] = "在庫限り"
        cleaned["note"] = clean_quote_free_text(str(cleaned.get("note") or "").replace("在庫限り", ""))
    if re.fullmatch(r"\d+\s*台", str(cleaned.get("note") or "").strip()):
        cleaned["note"] = ""
    return cleaned


def parse_quote_with_gemini(text: str, api_key: str) -> dict[str, object] | None:
    model = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    prompt = f"""
日本語の見積依頼テキストから、見積フォーム入力用の項目をJSONだけで抽出してください。
不明な項目は空文字にしてください。日付は今日={today().strftime('%Y-%m-%d')}を基準にYYYY-MM-DDへ変換してください。
回答期限、運賃、支払条件、FAX済みは抽出しないでください。

キー:
customer_query, customer_code, staff_name, quote_date, product_query, product_code, quantity, retail_price, wholesale_price, discounted_price, discount_name, discount_rate, stock_status, production_date, production_text, ship_date, ship_text, note

ルール:
- 赤字や追記がある場合は note、生産日、出荷予定日を優先してください。
- 「8月以降で回答していい？」のような内容は note に入れてください。
- 数量/台数がなければ1。
- 金額はカンマなしの数値。

テキスト:
{text}
""".strip()
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"responseMimeType": "application/json"},
    }
    return request_gemini_json(endpoint, body, api_key, timeout=20)


def parse_quote_image_with_gemini(image_bytes: bytes, filename: str, api_key: str) -> dict[str, object] | None:
    model = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    mime_type = mimetypes.guess_type(filename)[0] or "image/jpeg"
    prompt = f"""
見積書または見積依頼書の画像を読み取り、見積フォーム入力用の項目をJSONだけで抽出してください。
読み取れない項目は空文字にしてください。日付は今日={today().strftime('%Y-%m-%d')}を基準にYYYY-MM-DDへ正規化してください。
回答期限、運賃、支払条件、FAX済みは抽出しないでください。

キー:
document_type, visible_text, customer_query, customer_code, staff_name, quote_date, product_query, product_code, quantity, retail_price, wholesale_price, discounted_price, discount_name, discount_rate, stock_status, production_date, production_text, ship_date, ship_text, note

読み取りルール:
- 「見積依頼書」は依頼元の会社名、型式/品名、数量、納期、連絡事項を読む。
- 「見積書」は得意先名、商品コード、商品名、単価、在庫、備考の生産日/出荷日を読む。
- 伊藤産業機械は customer_code 61105、大竹産業は customer_code 61323。
- 赤字/紫字/手書き追記は重要。例: 「表は7/23(木)に記入しました」「8月以降で回答していい？」。
- 「生産日」「出荷日」「次回生産」「出荷予定」は production_date/ship_date または production_text/ship_text に入れる。
- 月だけの場合は production_text/ship_text に「8月」「2026年8月以降」のように入れる。
- 数量/台数がなければ1。
- 金額はカンマなしの数値。
""".strip()
    body = {
        "contents": [
            {
                "parts": [
                    {"text": prompt},
                    {
                        "inlineData": {
                            "mimeType": mime_type,
                            "data": base64.b64encode(image_bytes).decode("ascii"),
                        }
                    },
                ]
            }
        ],
        "generationConfig": {"responseMimeType": "application/json"},
    }
    return request_gemini_json(endpoint, body, api_key, timeout=25)


def resolve_quote_fields(parsed: dict[str, object]) -> dict[str, object]:
    result = coerce_parsed_dict(parsed)
    blob = " ".join(str(value or "") for value in result.values())
    if "大竹産業" in blob and not result.get("customer_code"):
        result["customer_code"] = "61323"
        result["customer_query"] = "大竹産業"
    if "伊藤産業機械" in blob and not result.get("customer_code"):
        result["customer_code"] = "61105"
        result["customer_query"] = "伊藤産業機械"
    if not result.get("customer_code") and any(term in blob for term in ["伊藤", "AM65B"]):
        result["customer_code"] = "61105"
        result["customer_query"] = "伊藤産業機械"
    if not result.get("stock_status"):
        for status in ["在庫限り", "生産待ち", "在庫有り", "有り", "あり", "無し", "なし"]:
            if status in blob:
                result["stock_status"] = "有り" if status == "あり" else ("無し" if status == "なし" else status)
                break
    if not result.get("note"):
        note_parts = []
        for pattern in [r"8月以降[^、。\n]*", r"表は[^、。\n]*記入[^、。\n]*"]:
            m = re.search(pattern, blob)
            if m:
                note_parts.append(m.group(0))
        if note_parts:
            result["note"] = " / ".join(note_parts)
    result = clean_quote_auxiliary_fields(result)
    if not result.get("quote_date"):
        result["quote_date"] = today().strftime("%Y-%m-%d")
    try:
        result["quantity"] = int(result.get("quantity") or 1)
    except (TypeError, ValueError):
        result["quantity"] = 1

    exact_customer = row_by_code(MASTER["customers"], result.get("customer_code"))
    if exact_customer:
        customer_candidates = [exact_customer]
    else:
        customer_candidates = ranked_matches(str(result.get("customer_query", "")), MASTER["customers"], ("name", "kana", "code"))
    customer = customer_candidates[0] if customer_candidates else None

    exact_product = row_by_code(MASTER["products"], result.get("product_code"))
    if exact_product:
        product_candidates = [exact_product]
    else:
        product_search_text = " ".join(str(result.get(key, "") or "") for key in ["product_code", "product_query", "visible_text"])
        token_candidates = []
        for token in model_tokens_from_text(product_search_text):
            token_candidates = ranked_matches(token, MASTER["products"], ("name", "code"))
            if token_candidates and float(token_candidates[0].get("score", 0)) >= 0.9:
                break
        product_candidates = token_candidates or ranked_matches(str(result.get("product_query", "")), MASTER["products"], ("name", "code"))
    product = product_candidates[0] if product_candidates else None

    retail_price = numeric_value(result.get("retail_price"))
    if not retail_price and product:
        retail_price = numeric_value(product.get("price"))
    wholesale_price = numeric_value(result.get("wholesale_price"))
    if not wholesale_price and retail_price and customer:
        rate = customer.get("rate")
        try:
            wholesale_price = int(round(retail_price * float(rate)))
        except (TypeError, ValueError):
            wholesale_price = 0
    discounted_price = numeric_value(result.get("discounted_price"))
    discount_rate = str(result.get("discount_rate") or "")
    if not discounted_price and wholesale_price and discount_rate:
        try:
            discounted_price = int(round(wholesale_price * (1 - float(discount_rate) / 100)))
        except ValueError:
            discounted_price = 0

    return {
        **result,
        "customer": customer,
        "customer_candidates": customer_candidates,
        "customer_name": customer.get("name", "") if customer else str(result.get("customer_query", "") or ""),
        "customer_code": customer.get("code", "") if customer else str(result.get("customer_code", "") or ""),
        "customer_needs_confirmation": needs_confirmation(customer_candidates),
        "product": product,
        "product_candidates": product_candidates,
        "product_name": product.get("name", "") if product else str(result.get("product_query", "") or ""),
        "product_code": product.get("code", "") if product else str(result.get("product_code", "") or ""),
        "product_needs_confirmation": needs_confirmation(product_candidates),
        "retail_price": retail_price,
        "wholesale_price": wholesale_price,
        "discounted_price": discounted_price,
        "discount_rate": discount_rate,
        "quantity": result.get("quantity") or 1,
        "stock_status": result.get("stock_status") or "",
        "production_date": result.get("production_date") or "",
        "production_text": result.get("production_text") or "",
        "ship_date": result.get("ship_date") or "",
        "ship_text": result.get("ship_text") or "",
        "note": result.get("note") or "",
    }


def style_quote_sheet(ws) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Yu Gothic", size=10)
    for row in ws.iter_rows(min_row=11, max_row=13, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
    for cell in ws[11]:
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.font = Font(name="Yu Gothic", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {"A": 14, "B": 20, "C": 12, "D": 14, "E": 14, "F": 12, "G": 13, "H": 24}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def save_quote_estimate_excel(fields: dict[str, object]) -> Path:
    fields = resolve_quote_fields(fields)
    output_path = quote_output_path(fields, "見積書")
    wb = Workbook()
    ws = wb.active
    ws.title = "見積書"
    customer = fields.get("customer") or {}
    product = fields.get("product") or {}

    ws.merge_cells("A1:H1")
    ws["A1"] = "見　積　書"
    ws["A1"].font = Font(name="Yu Gothic", size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A3"] = customer.get("code", "")
    ws["B4"] = fields.get("customer_name") or customer.get("name", "")
    ws["D4"] = "御中"
    ws["G2"] = "作成日"
    ws["H2"] = fields.get("quote_date") or today().strftime("%Y-%m-%d")
    ws["F4"] = "㈱ オーレック関東営業G"
    ws["F5"] = "TEL: 0480-50-9020"
    ws["F6"] = "FAX: 0480-87-3009"
    ws["B7"] = "毎度格別のお引き立てを賜りお礼申し上げます。"
    ws["B8"] = "下記の通りお見積り申し上げます。"

    headers = ["商品コード", "商品名", "小売単価", "仕切単価", "値引後単価", "台数", "在庫", "備考"]
    for col, header in enumerate(headers, 1):
        ws.cell(11, col).value = header
    ws["A12"] = fields.get("product_code") or product.get("code", "")
    ws["B12"] = fields.get("product_name") or product.get("name", "")
    ws["C12"] = numeric_value(fields.get("retail_price"))
    ws["D12"] = numeric_value(fields.get("wholesale_price"))
    ws["E12"] = numeric_value(fields.get("discounted_price"))
    ws["F12"] = int(fields.get("quantity") or 1)
    ws["G12"] = fields.get("stock_status", "")
    note_parts = []
    if fields.get("production_date") or fields.get("production_text"):
        note_parts.append(f"生産日 {fields.get('production_date') or fields.get('production_text')}")
    if fields.get("ship_date") or fields.get("ship_text"):
        note_parts.append(f"出荷日 {fields.get('ship_date') or fields.get('ship_text')}")
    if fields.get("discount_name") or fields.get("discount_rate"):
        note_parts.append(f"{fields.get('discount_name') or '値引'} {fields.get('discount_rate')}%".strip())
    if fields.get("note"):
        note_parts.append(str(fields.get("note")))
    ws["H12"] = "\n".join(part for part in note_parts if part)

    for cell in ["C12", "D12", "E12"]:
        ws[cell].number_format = "#,##0"
    ws["B16"] = "尚、生産予定日は諸事情により変動する場合がございますのでご了承願います。"
    ws["B18"] = "見積り時点での商品確保はできかねます。"
    style_quote_sheet(ws)
    wb.save(output_path)
    return output_path


def save_quote_request_excel(fields: dict[str, object]) -> Path:
    fields = resolve_quote_fields(fields)
    output_path = quote_output_path(fields, "見積依頼")
    wb = Workbook()
    ws = wb.active
    ws.title = "見積依頼"
    customer = fields.get("customer") or {}
    product = fields.get("product") or {}

    ws.merge_cells("A1:H1")
    ws["A1"] = "見 積 依 頼 控 え"
    ws["A1"].font = Font(name="Yu Gothic", size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A3"] = "得意先"
    ws["B3"] = fields.get("customer_name") or customer.get("name", "")
    ws["F3"] = "依頼日"
    ws["G3"] = fields.get("quote_date") or today().strftime("%Y-%m-%d")
    ws["A4"] = "担当者"
    ws["B4"] = fields.get("staff_name", "")

    headers = ["商品コード", "商品名", "小売単価", "仕切単価", "値引後単価", "数量", "在庫", "備考"]
    for col, header in enumerate(headers, 1):
        ws.cell(7, col).value = header
    ws["A8"] = fields.get("product_code") or product.get("code", "")
    ws["B8"] = fields.get("product_name") or product.get("name", "")
    ws["C8"] = numeric_value(fields.get("retail_price"))
    ws["D8"] = numeric_value(fields.get("wholesale_price"))
    ws["E8"] = numeric_value(fields.get("discounted_price"))
    ws["F8"] = int(fields.get("quantity") or 1)
    ws["G8"] = fields.get("stock_status", "")
    ws["H8"] = fields.get("note", "")
    ws["A11"] = "生産日"
    ws["B11"] = fields.get("production_date") or fields.get("production_text") or ""
    ws["A12"] = "出荷予定日"
    ws["B12"] = fields.get("ship_date") or fields.get("ship_text") or ""
    style_quote_sheet(ws)
    wb.save(output_path)
    return output_path


def save_excel(fields: dict[str, object]) -> Path:
    fields = resolve_default_dates(fields)
    output_path = output_excel_path(fields)
    shutil.copy2(SOURCE_BOOK, output_path)
    wb = load_workbook(output_path)
    ws = wb["引合書+値引"]

    customer = fields.get("customer") or {}
    delivery = fields.get("delivery") or {}
    product = fields.get("product") or {}

    ws["A4"] = customer.get("code", "")
    ws["A6"] = delivery.get("code", customer.get("code", ""))
    ws["I4"] = fields.get("order_date") or today().strftime("%Y-%m-%d")
    ws["H5"] = fields.get("production_date") or fields.get("production_text") or "在庫"
    ws["I6"] = fields.get("ship_date") or fields.get("ship_text") or ""
    ws["B17"] = normalize_warehouse(fields.get("warehouse")) or "011"
    ws["L7"] = fields.get("discount_name") or "音声入力値引"
    ws["L8"] = fields.get("discount_method") or "外掛"
    ws["N8"] = float(fields.get("discount_rate") or 0)
    ws["A11"] = product.get("code", "")
    ws["E11"] = int(fields.get("quantity") or 1)
    ws["I11"] = fields.get("order_no", "")

    ws["I4"].number_format = "yyyy-mm-dd"
    ws["H5"].number_format = "yyyy-mm-dd" if fields.get("production_date") else "@"
    ws["I6"].number_format = "yyyy-mm-dd" if fields.get("ship_date") else "@"
    ws["B17"].number_format = "@"
    ws["N8"].number_format = "0.0"

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.save(output_path)
    return output_path


def create_download_token(path: Path) -> str:
    token = secrets.token_urlsafe(32)
    expires_at = time.time() + int(os.getenv("DOWNLOAD_TOKEN_MINUTES", "60")) * 60
    DOWNLOAD_TOKENS[token] = (path, expires_at)
    return token


HTML = """
<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>引合書 音声入力フォーム</title>
  <style>
    body { font-family: system-ui, -apple-system, "Segoe UI", sans-serif; margin: 0; background: #f5f7fb; color: #172033; }
    main { max-width: 1100px; margin: 0 auto; padding: 28px; }
    h1 { font-size: 26px; margin: 0 0 18px; }
    .panel { background: white; border: 1px solid #d9e2ef; border-radius: 8px; padding: 18px; margin-bottom: 16px; }
    textarea { width: 100%; min-height: 96px; font-size: 17px; padding: 12px; box-sizing: border-box; }
    button { border: 0; border-radius: 6px; padding: 11px 16px; font-size: 15px; cursor: pointer; background: #1f4e79; color: white; }
    button.secondary { background: #60758b; }
    button.save { background: #2f7d32; }
    button:disabled { opacity: .55; cursor: not-allowed; }
    .row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; }
    label { display: block; font-size: 13px; color: #435064; margin-bottom: 5px; }
    input, select { width: 100%; box-sizing: border-box; padding: 9px; border: 1px solid #c8d3df; border-radius: 5px; font-size: 15px; }
    .actions { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 12px; }
    .status { margin-top: 10px; color: #435064; white-space: pre-wrap; }
    .match { font-size: 12px; color: #5f6d7c; margin-top: 4px; }
    .confirm { color: #9a4d00; font-weight: 700; }
    .fileline { display: flex; gap: 10px; flex-wrap: wrap; align-items: end; }
    .fileline input { max-width: 520px; }
    a { color: #1f4e79; font-weight: 600; }
    @media (max-width: 800px) { .row { grid-template-columns: 1fr; } main { padding: 16px; } }
  </style>
</head>
<body>
<main>
  <h1>引合書 音声入力フォーム</h1>
  <p><a href="/quote">見積り依頼フォームへ</a></p>
  <section class="panel">
    <label for="voiceText">音声または手入力</label>
    <textarea id="voiceText">得意先 良栄社、受注日 今日、出荷希望日 明後日、倉庫011、値引き外掛け、値引き率3%、製品SP853A</textarea>
    <div class="actions">
      <button id="listenBtn">音声入力開始</button>
      <button class="secondary" id="parseBtn">内容をフォームへ反映</button>
    </div>
    <div class="status" id="speechStatus"></div>
  </section>

  <section class="panel">
    <label for="orderImage">注文書写真から読み取り</label>
    <div class="fileline">
      <input id="orderImage" type="file" accept=".jpg,.jpeg,.png,.webp,.pdf">
      <button class="secondary" id="imageParseBtn">写真を読み取ってフォームへ反映</button>
    </div>
    <div class="status" id="imageStatus">Gemini API の接続状態を確認中...</div>
  </section>

  <section class="panel">
    <div class="row">
      <div><label>得意先名</label><input id="customer_name"><select id="customer_candidates"></select><div class="match" id="customer_match"></div></div>
      <div><label>得意先コード</label><input id="customer_code"></div>
      <div><label>納入先名</label><input id="delivery_name"><select id="delivery_candidates"></select><div class="match" id="delivery_match"></div></div>
      <div><label>納入先コード</label><input id="delivery_code"></div>
      <div><label>受注日</label><input id="order_date" type="date"></div>
      <div><label>生産日</label><input id="production_date" type="date"><div class="match">未指定の場合は帳票に「在庫」と入ります</div></div>
      <div><label>出荷希望日</label><input id="ship_date" type="date"></div>
      <div><label>倉庫</label><select id="warehouse"><option>011</option><option>031</option></select></div>
      <div><label>値引方法</label><select id="discount_method"><option>外掛</option><option>内掛</option></select></div>
      <div><label>値引率(%)</label><input id="discount_rate" type="number" step="0.1"></div>
      <div><label>製品名</label><input id="product_name"><select id="product_candidates"></select><div class="match" id="product_match"></div></div>
      <div><label>商品コード</label><input id="product_code"></div>
      <div><label>数量</label><input id="quantity" type="number" value="1"></div>
    </div>
    <div class="actions">
      <button class="save" id="saveBtn">新規Excelとして保存</button>
    </div>
    <div class="status" id="saveStatus"></div>
  </section>
</main>
<script>
const $ = (id) => document.getElementById(id);
let currentFields = {};

function setStatus(id, text) { $(id).textContent = text; }

async function loadStatus() {
  try {
    const res = await fetch("/api/status");
    const data = await res.json();
    if (data.gemini_configured) {
      setStatus("imageStatus", `Gemini API 接続設定あり (${data.gemini_model})。注文書写真を読み取れます。`);
    } else {
      setStatus("imageStatus", "画像読み取りには GEMINI_API_KEY が必要です。");
    }
  } catch {
    setStatus("imageStatus", "Gemini API の接続状態を確認できませんでした。");
  }
}

loadStatus();

function optionLabel(item, type) {
  if (!item) return "";
  const score = item.score ? ` / 一致度 ${item.score}` : "";
  const place = item.address ? ` / ${item.address}` : "";
  const price = item.price ? ` / ${item.price}` : "";
  return `${item.name || ""} / ${item.code || ""}${place}${price}${score}`;
}

function fillSelect(selectId, candidates, selectedCode) {
  const select = $(selectId);
  select.innerHTML = "";
  if (!candidates || candidates.length === 0) {
    const opt = document.createElement("option");
    opt.value = "";
    opt.textContent = "候補なし";
    select.appendChild(opt);
    return;
  }
  candidates.forEach((item, index) => {
    const opt = document.createElement("option");
    opt.value = String(index);
    opt.textContent = optionLabel(item);
    if (selectedCode && String(item.code) === String(selectedCode)) opt.selected = true;
    select.appendChild(opt);
  });
}

function applyCandidate(kind, index) {
  const list = currentFields[`${kind}_andidates`] || currentFields[`${kind}_candidates`] || [];
  const item = list[Number(index)];
  if (!item) return;
  if (kind === "customer") {
    currentFields.customer = item;
    $("customer_name").value = item.name || "";
    $("customer_code").value = item.code || "";
  } else if (kind === "delivery") {
    currentFields.delivery = item;
    $("delivery_name").value = item.name || "";
    $("delivery_code").value = item.code || "";
  } else if (kind === "product") {
    currentFields.product = item;
    $("product_name").value = item.name || "";
    $("product_code").value = item.code || "";
  }
}

function applyFields(data, statusId) {
  currentFields = data.fields;
  const f = currentFields;
  $("customer_name").value = f.customer?.name || f.customer_query || "";
  $("customer_code").value = f.customer?.code || "";
  $("delivery_name").value = f.delivery?.name || f.delivery_query || "";
  $("delivery_code").value = f.delivery?.code || "";
  $("order_date").value = f.order_date || "";
  $("production_date").value = f.production_date || "";
  $("ship_date").value = f.ship_date || "";
  $("warehouse").value = f.warehouse || "011";
  $("discount_method").value = f.discount_method || "外掛";
  $("discount_rate").value = f.discount_rate || "";
  $("product_name").value = f.product?.name || f.product_query || "";
  $("product_code").value = f.product?.code || "";
  $("quantity").value = f.quantity || 1;

  fillSelect("customer_candidates", f.customer_candidates, f.customer?.code);
  fillSelect("delivery_candidates", f.delivery_candidates, f.delivery?.code);
  fillSelect("product_candidates", f.product_candidates, f.product?.code);

  $("customer_match").innerHTML = f.customer
    ? `${f.customer_needs_confirmation ? '<span class="confirm">候補確認が必要です。</span> ' : ''}一致度 ${f.customer.score}`
    : "候補なし";
  $("delivery_match").innerHTML = f.delivery
    ? `${f.delivery_needs_confirmation ? '<span class="confirm">候補確認が必要です。</span> ' : ''}一致度 ${f.delivery.score}`
    : "候補なし";
  $("product_match").innerHTML = f.product
    ? `${f.product_needs_confirmation ? '<span class="confirm">候補確認が必要です。</span> ' : ''}一致度 ${f.product.score}`
    : "候補なし";
  const warnings = [];
  if (f.customer_needs_confirmation) warnings.push("得意先");
  if (f.delivery_needs_confirmation) warnings.push("納入先");
  if (f.product_needs_confirmation) warnings.push("製品");
  setStatus(statusId, `解析完了 (${data.parser})${warnings.length ? "\\n確認してください: " + warnings.join("、") : ""}`);
}

$("customer_candidates").onchange = (e) => applyCandidate("customer", e.target.value);
$("delivery_candidates").onchange = (e) => applyCandidate("delivery", e.target.value);
$("product_candidates").onchange = (e) => applyCandidate("product", e.target.value);

function debounce(fn, wait) {
  let timer;
  return (...args) => {
    clearTimeout(timer);
    timer = setTimeout(() => fn(...args), wait);
  };
}

function addBusinessDaysIso(isoDate, days) {
  if (!isoDate) return "";
  const date = new Date(`${isoDate}T00:00:00`);
  if (Number.isNaN(date.getTime())) return "";
  let added = 0;
  while (added < days) {
    date.setDate(date.getDate() + 1);
    const day = date.getDay();
    if (day !== 0 && day !== 6) added += 1;
  }
  return date.toISOString().slice(0, 10);
}

async function suggest(kind, query) {
  if (!query || query.trim().length < 1) return;
  const res = await fetch(`/api/suggest?kind=${encodeURIComponent(kind)}&q=${encodeURIComponent(query)}`);
  const data = await res.json();
  if (!res.ok) return;
  const candidatesKey = `${kind}_candidates`;
  currentFields[candidatesKey] = data.candidates || [];
  currentFields[`${kind}_needs_confirmation`] = data.needs_confirmation;

  const selectId = `${kind}_candidates`;
  fillSelect(selectId, data.candidates, data.candidates?.[0]?.code);
  const first = data.candidates?.[0];
  const labelId = `${kind}_match`;
  if (!first) {
    $(labelId).textContent = "候補なし";
    return;
  }
  const confirmText = data.needs_confirmation ? "候補確認が必要です。 " : "";
  $(labelId).innerHTML = `${confirmText}一致度 ${first.score}`;

  if (Number(first.score) >= 1.0) {
    currentFields[kind] = first;
    if (kind === "customer") {
      $("customer_name").value = first.name || query;
      $("customer_code").value = first.code || "";
    } else if (kind === "delivery") {
      $("delivery_name").value = first.name || query;
      $("delivery_code").value = first.code || "";
    } else if (kind === "product") {
      $("product_name").value = first.name || query;
      $("product_code").value = first.code || "";
    }
  }
}

$("customer_name").addEventListener("input", debounce((e) => suggest("customer", e.target.value), 250));
$("delivery_name").addEventListener("input", debounce((e) => suggest("delivery", e.target.value), 250));
$("product_name").addEventListener("input", debounce((e) => suggest("product", e.target.value), 250));
$("production_date").addEventListener("change", (e) => {
  if (!$("ship_date").value) {
    $("ship_date").value = addBusinessDaysIso(e.target.value, 2);
  }
});

$("listenBtn").onclick = () => {
  const SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;
  if (!SpeechRecognition) {
    setStatus("speechStatus", "このブラウザはWeb Speech APIに未対応です。ChromeまたはEdgeで開いてください。");
    return;
  }
  const maxListenMs = 60 * 1000;
  const silenceMs = 20 * 1000;
  const rec = new SpeechRecognition();
  rec.lang = "ja-JP";
  rec.interimResults = true;
  rec.continuous = true;
  $("listenBtn").disabled = true;
  setStatus("speechStatus", "聞き取り中... 最大60秒、20秒無音で自動停止します。");
  let stoppedByTimer = "";
  let stoppedByError = false;
  let maxTimer = null;
  let silenceTimer = null;
  const stopListening = (reason) => {
    stoppedByTimer = reason;
    try { rec.stop(); } catch (_) {}
  };
  const resetSilenceTimer = () => {
    clearTimeout(silenceTimer);
    silenceTimer = setTimeout(() => stopListening("20秒無音だったため停止しました。"), silenceMs);
  };
  maxTimer = setTimeout(() => stopListening("60秒に達したため停止しました。"), maxListenMs);
  resetSilenceTimer();
  rec.onresult = (event) => {
    resetSilenceTimer();
    let text = "";
    for (let i = 0; i < event.results.length; i++) text += event.results[i][0].transcript;
    $("voiceText").value = text;
  };
  rec.onspeechstart = resetSilenceTimer;
  rec.onsoundstart = resetSilenceTimer;
  rec.onerror = (event) => {
    stoppedByError = true;
    clearTimeout(maxTimer);
    clearTimeout(silenceTimer);
    $("listenBtn").disabled = false;
    setStatus("speechStatus", "音声入力エラー: " + event.error);
  };
  rec.onend = () => {
    clearTimeout(maxTimer);
    clearTimeout(silenceTimer);
    $("listenBtn").disabled = false;
    if (stoppedByError) return;
    const message = stoppedByTimer || "聞き取り完了。";
    setStatus("speechStatus", `${message} 必要なら文章を修正してから反映してください。`);
  };
  rec.start();
};

$("parseBtn").onclick = async () => {
  setStatus("speechStatus", "解析中...");
  const res = await fetch("/api/parse", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({ text: $("voiceText").value })
  });
  const data = await res.json();
  if (!res.ok) {
    setStatus("speechStatus", data.error || "解析に失敗しました。");
    return;
  }
  applyFields(data, "speechStatus");
};

$("imageParseBtn").onclick = async () => {
  const file = $("orderImage").files[0];
  if (!file) {
    setStatus("imageStatus", "注文書写真を選択してください。");
    return;
  }
  setStatus("imageStatus", "画像を読み取り中...");
  const form = new FormData();
  form.append("image", file);
  const res = await fetch("/api/parse-image", { method: "POST", body: form });
  const data = await res.json();
  if (!res.ok) {
    setStatus("imageStatus", data.error || "画像読み取りに失敗しました。");
    return;
  }
  applyFields(data, "imageStatus");
};

$("saveBtn").onclick = async () => {
  const fields = {
    ...currentFields,
    customer: { ...(currentFields.customer || {}), name: $("customer_name").value, code: $("customer_code").value },
    delivery: { ...(currentFields.delivery || {}), name: $("delivery_name").value, code: $("delivery_code").value },
    product: { ...(currentFields.product || {}), name: $("product_name").value, code: $("product_code").value },
    order_date: $("order_date").value,
    production_date: $("production_date").value,
    ship_date: $("ship_date").value,
    warehouse: $("warehouse").value,
    discount_method: $("discount_method").value,
    discount_rate: $("discount_rate").value,
    quantity: $("quantity").value
  };
  setStatus("saveStatus", "Excel保存中...");
  const res = await fetch("/api/save", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({ fields })
  });
  const data = await res.json();
  if (!res.ok) {
    setStatus("saveStatus", data.error || "保存に失敗しました。");
    return;
  }
  setStatus("saveStatus", "保存しました: " + data.path);
  const a = document.createElement("a");
  a.href = data.download_url;
  a.textContent = "保存したExcelをダウンロード";
  a.style.display = "block";
  a.style.marginTop = "8px";
  $("saveStatus").appendChild(a);
};
</script>
</body>
</html>
"""


QUOTE_HTML = """
<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>見積り依頼フォーム</title>
  <style>
    body { font-family: system-ui, -apple-system, "Segoe UI", sans-serif; margin: 0; background: #f5f7fb; color: #172033; }
    main { max-width: 1100px; margin: 0 auto; padding: 28px; }
    h1 { font-size: 26px; margin: 0 0 18px; }
    .panel { background: white; border: 1px solid #d9e2ef; border-radius: 8px; padding: 18px; margin-bottom: 16px; }
    textarea { width: 100%; min-height: 96px; font-size: 17px; padding: 12px; box-sizing: border-box; }
    button { border: 0; border-radius: 6px; padding: 11px 16px; font-size: 15px; cursor: pointer; background: #1f4e79; color: white; }
    button.secondary { background: #60758b; }
    button.save { background: #2f7d32; }
    button:disabled { opacity: .55; cursor: not-allowed; }
    .row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; }
    label { display: block; font-size: 13px; color: #435064; margin-bottom: 5px; }
    input, select { width: 100%; box-sizing: border-box; padding: 9px; border: 1px solid #c8d3df; border-radius: 5px; font-size: 15px; }
    .actions { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 12px; }
    .status { margin-top: 10px; color: #435064; white-space: pre-wrap; }
    .match { font-size: 12px; color: #5f6d7c; margin-top: 4px; }
    .confirm { color: #9a4d00; font-weight: 700; }
    .fileline { display: flex; gap: 10px; flex-wrap: wrap; align-items: end; }
    a { color: #1f4e79; font-weight: 600; }
    @media (max-width: 800px) { .row { grid-template-columns: 1fr; } main { padding: 16px; } }
  </style>
</head>
<body>
<main>
  <h1>見積り依頼フォーム</h1>
  <p><a href="/">引合書フォームへ戻る</a></p>

  <section class="panel">
    <label for="quoteText">音声または手入力</label>
    <textarea id="quoteText">得意先 大竹産業、担当 矢村、製品 RCSP540、数量1台、生産待ち、出荷予定 7月13日</textarea>
    <div class="actions">
      <button id="quoteListenBtn">音声入力開始</button>
      <button class="secondary" id="quoteParseBtn">内容をフォームへ反映</button>
    </div>
    <div class="status" id="quoteSpeechStatus"></div>
  </section>

  <section class="panel">
    <label for="quoteImage">見積依頼写真から読み取り</label>
    <div class="fileline">
      <input id="quoteImage" type="file" accept=".jpg,.jpeg,.png,.webp,.pdf">
      <button class="secondary" id="quoteImageParseBtn">写真を読み取ってフォームへ反映</button>
    </div>
    <div class="status" id="quoteImageStatus">Gemini API の接続状態を確認中...</div>
  </section>

  <section class="panel">
    <div class="row">
      <div><label>得意先名</label><input id="quote_customer_name"><select id="quote_customer_candidates"></select><div class="match" id="quote_customer_match"></div></div>
      <div><label>得意先コード</label><input id="quote_customer_code"></div>
      <div><label>担当者</label><input id="quote_staff_name"></div>
      <div><label>依頼日/作成日</label><input id="quote_date" type="date"></div>
      <div><label>商品名/機種名</label><input id="quote_product_name"><select id="quote_product_candidates"></select><div class="match" id="quote_product_match"></div></div>
      <div><label>商品コード</label><input id="quote_product_code"></div>
      <div><label>数量/台数</label><input id="quote_quantity" type="number" value="1"></div>
      <div><label>在庫状態</label><input id="quote_stock_status"></div>
      <div><label>小売単価</label><input id="quote_retail_price" type="number"></div>
      <div><label>仕切単価</label><input id="quote_wholesale_price" type="number"></div>
      <div><label>値引後単価</label><input id="quote_discounted_price" type="number"></div>
      <div><label>値引名</label><input id="quote_discount_name"></div>
      <div><label>値引率(%)</label><input id="quote_discount_rate" type="number" step="0.1"></div>
      <div><label>生産日</label><input id="quote_production_date" type="date"></div>
      <div><label>出荷予定日</label><input id="quote_ship_date" type="date"></div>
      <div><label>備考</label><input id="quote_note"></div>
    </div>
    <div class="actions">
      <button class="save" id="quoteEstimateSaveBtn">見積書Excelとして保存</button>
      <button class="save" id="quoteRequestSaveBtn">見積依頼控えExcelとして保存</button>
    </div>
    <div class="status" id="quoteSaveStatus"></div>
  </section>
</main>
<script>
const $ = (id) => document.getElementById(id);
let quoteFields = {};

function setStatus(id, message) { $(id).textContent = message || ""; }

async function statusCheck() {
  try {
    const res = await fetch("/api/status");
    const data = await res.json();
    setStatus("quoteImageStatus", data.gemini_configured ? `Gemini API 接続設定あり (${data.gemini_model})。` : "画像読み取りには GEMINI_API_KEY が必要です。");
  } catch (_) {
    setStatus("quoteImageStatus", "Gemini API の接続状態を確認できませんでした。");
  }
}
statusCheck();

function fillSelect(selectId, candidates, selectedCode) {
  const select = $(selectId);
  select.innerHTML = "";
  (candidates || []).forEach((item, index) => {
    const opt = document.createElement("option");
    opt.value = index;
    opt.selected = selectedCode && String(item.code) === String(selectedCode);
    opt.textContent = `${item.name || ""} / ${item.code || ""} / 一致度 ${item.score ?? ""}`;
    select.appendChild(opt);
  });
}

function applyCandidate(kind, index) {
  const list = quoteFields[`${kind}_candidates`] || [];
  const item = list[Number(index)];
  if (!item) return;
  if (kind === "customer") {
    quoteFields.customer = item;
    $("quote_customer_name").value = item.name || "";
    $("quote_customer_code").value = item.code || "";
  } else if (kind === "product") {
    quoteFields.product = item;
    $("quote_product_name").value = item.name || "";
    $("quote_product_code").value = item.code || "";
    if (!$("quote_retail_price").value && item.price) $("quote_retail_price").value = item.price;
  }
}

function applyQuoteFields(data, statusId) {
  quoteFields = data.fields || {};
  const f = quoteFields;
  $("quote_customer_name").value = f.customer?.name || f.customer_query || "";
  $("quote_customer_code").value = f.customer?.code || f.customer_code || "";
  $("quote_staff_name").value = f.staff_name || "";
  $("quote_date").value = f.quote_date || "";
  $("quote_product_name").value = f.product?.name || f.product_query || "";
  $("quote_product_code").value = f.product?.code || f.product_code || "";
  $("quote_quantity").value = f.quantity || 1;
  $("quote_stock_status").value = f.stock_status || "";
  $("quote_retail_price").value = f.retail_price || "";
  $("quote_wholesale_price").value = f.wholesale_price || "";
  $("quote_discounted_price").value = f.discounted_price || "";
  $("quote_discount_name").value = f.discount_name || "";
  $("quote_discount_rate").value = f.discount_rate || "";
  $("quote_production_date").value = f.production_date || "";
  $("quote_ship_date").value = f.ship_date || "";
  $("quote_note").value = f.note || f.production_text || f.ship_text || "";
  fillSelect("quote_customer_candidates", f.customer_candidates, f.customer?.code);
  fillSelect("quote_product_candidates", f.product_candidates, f.product?.code);
  $("quote_customer_match").innerHTML = f.customer ? `${f.customer_needs_confirmation ? '<span class="confirm">候補確認が必要です。</span> ' : ''}一致度 ${f.customer.score}` : "候補なし";
  $("quote_product_match").innerHTML = f.product ? `${f.product_needs_confirmation ? '<span class="confirm">候補確認が必要です。</span> ' : ''}一致度 ${f.product.score}` : "候補なし";
  const warnings = [];
  if (f.customer_needs_confirmation) warnings.push("得意先");
  if (f.product_needs_confirmation) warnings.push("製品");
  setStatus(statusId, `解析完了 (${data.parser})${warnings.length ? "\\n確認してください: " + warnings.join("、") : ""}`);
}

$("quote_customer_candidates").onchange = (e) => applyCandidate("customer", e.target.value);
$("quote_product_candidates").onchange = (e) => applyCandidate("product", e.target.value);

function debounce(fn, wait) {
  let timer;
  return (...args) => {
    clearTimeout(timer);
    timer = setTimeout(() => fn(...args), wait);
  };
}

async function suggest(kind, query) {
  if (!query || query.trim().length < 1) return;
  const res = await fetch(`/api/suggest?kind=${encodeURIComponent(kind)}&q=${encodeURIComponent(query)}`);
  const data = await res.json();
  if (!res.ok) return;
  quoteFields[`${kind}_candidates`] = data.candidates || [];
  quoteFields[`${kind}_needs_confirmation`] = data.needs_confirmation;
  fillSelect(`quote_${kind}_candidates`, data.candidates, data.candidates?.[0]?.code);
  const first = data.candidates?.[0];
  const label = $(`quote_${kind}_match`);
  label.innerHTML = first ? `${data.needs_confirmation ? '<span class="confirm">候補確認が必要です。</span> ' : ''}一致度 ${first.score}` : "候補なし";
}

$("quote_customer_name").addEventListener("input", debounce((e) => suggest("customer", e.target.value), 250));
$("quote_product_name").addEventListener("input", debounce((e) => suggest("product", e.target.value), 250));

$("quoteListenBtn").onclick = () => {
  const SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;
  if (!SpeechRecognition) {
    setStatus("quoteSpeechStatus", "このブラウザはWeb Speech APIに未対応です。ChromeまたはEdgeで開いてください。");
    return;
  }
  const rec = new SpeechRecognition();
  rec.lang = "ja-JP";
  rec.interimResults = true;
  rec.continuous = true;
  $("quoteListenBtn").disabled = true;
  setStatus("quoteSpeechStatus", "聞き取り中... 最大60秒、20秒無音で自動停止します。");
  let stoppedByTimer = "";
  let stoppedByError = false;
  let maxTimer = null;
  let silenceTimer = null;
  const stopListening = (reason) => {
    stoppedByTimer = reason;
    try { rec.stop(); } catch (_) {}
  };
  const resetSilenceTimer = () => {
    clearTimeout(silenceTimer);
    silenceTimer = setTimeout(() => stopListening("20秒無音だったため停止しました。"), 20 * 1000);
  };
  maxTimer = setTimeout(() => stopListening("60秒に達したため停止しました。"), 60 * 1000);
  resetSilenceTimer();
  rec.onresult = (event) => {
    resetSilenceTimer();
    let text = "";
    for (let i = 0; i < event.results.length; i++) text += event.results[i][0].transcript;
    $("quoteText").value = text;
  };
  rec.onspeechstart = resetSilenceTimer;
  rec.onsoundstart = resetSilenceTimer;
  rec.onerror = (event) => {
    stoppedByError = true;
    clearTimeout(maxTimer);
    clearTimeout(silenceTimer);
    $("quoteListenBtn").disabled = false;
    setStatus("quoteSpeechStatus", "音声入力エラー: " + event.error);
  };
  rec.onend = () => {
    clearTimeout(maxTimer);
    clearTimeout(silenceTimer);
    $("quoteListenBtn").disabled = false;
    if (stoppedByError) return;
    setStatus("quoteSpeechStatus", `${stoppedByTimer || "聞き取り完了。"} 必要なら文章を修正してから反映してください。`);
  };
  rec.start();
};

$("quoteParseBtn").onclick = async () => {
  setStatus("quoteSpeechStatus", "解析中...");
  const res = await fetch("/api/quote/parse", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({ text: $("quoteText").value })
  });
  const data = await res.json();
  if (!res.ok) {
    setStatus("quoteSpeechStatus", data.error || "解析に失敗しました。");
    return;
  }
  applyQuoteFields(data, "quoteSpeechStatus");
};

$("quoteImageParseBtn").onclick = async () => {
  const file = $("quoteImage").files[0];
  if (!file) {
    setStatus("quoteImageStatus", "見積依頼写真を選択してください。");
    return;
  }
  setStatus("quoteImageStatus", "画像を読み取り中...");
  const form = new FormData();
  form.append("image", file);
  const res = await fetch("/api/quote/parse-image", { method: "POST", body: form });
  const data = await res.json();
  if (!res.ok) {
    setStatus("quoteImageStatus", data.error || "画像読み取りに失敗しました。");
    return;
  }
  applyQuoteFields(data, "quoteImageStatus");
};

function collectQuoteFields() {
  return {
    ...quoteFields,
    customer: { ...(quoteFields.customer || {}), name: $("quote_customer_name").value, code: $("quote_customer_code").value },
    product: { ...(quoteFields.product || {}), name: $("quote_product_name").value, code: $("quote_product_code").value },
    customer_name: $("quote_customer_name").value,
    customer_code: $("quote_customer_code").value,
    staff_name: $("quote_staff_name").value,
    quote_date: $("quote_date").value,
    product_name: $("quote_product_name").value,
    product_code: $("quote_product_code").value,
    quantity: $("quote_quantity").value,
    stock_status: $("quote_stock_status").value,
    retail_price: $("quote_retail_price").value,
    wholesale_price: $("quote_wholesale_price").value,
    discounted_price: $("quote_discounted_price").value,
    discount_name: $("quote_discount_name").value,
    discount_rate: $("quote_discount_rate").value,
    production_date: $("quote_production_date").value,
    ship_date: $("quote_ship_date").value,
    note: $("quote_note").value
  };
}

async function saveQuote(kind) {
  setStatus("quoteSaveStatus", "Excel保存中...");
  const res = await fetch(`/api/quote/save-${kind}`, {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({ fields: collectQuoteFields() })
  });
  const data = await res.json();
  if (!res.ok) {
    setStatus("quoteSaveStatus", data.error || "保存に失敗しました。");
    return;
  }
  setStatus("quoteSaveStatus", "保存しました: " + data.path);
  const a = document.createElement("a");
  a.href = data.download_url;
  a.textContent = "保存したExcelをダウンロード";
  a.style.display = "block";
  a.style.marginTop = "8px";
  $("quoteSaveStatus").appendChild(a);
}

$("quoteEstimateSaveBtn").onclick = () => saveQuote("estimate");
$("quoteRequestSaveBtn").onclick = () => saveQuote("request");
</script>
</body>
</html>
"""


@app.get("/")
def index():
    return HTML


@app.get("/quote")
def quote_index():
    return QUOTE_HTML


@app.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["Permissions-Policy"] = "camera=(), geolocation=(), payment=()"
    response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data:; connect-src 'self'; object-src 'none'; base-uri 'self'; frame-ancestors 'none'"
    if env_bool("ENABLE_HSTS", False):
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response


@app.post("/api/parse")
def api_parse():
    payload = json_payload()
    if payload is None:
        return jsonify({"error": "JSON形式が正しくありません。"}), 400
    text = str(payload.get("text", "")).strip()
    if not text:
        return jsonify({"error": "音声テキストが空です。"}), 400
    if len(text) > max_text_chars():
        audit("parse_text_too_long", length=len(text), limit=max_text_chars())
        return jsonify({"error": f"入力が長すぎます。{max_text_chars()}文字以内にしてください。"}), 400

    api_key = os.getenv("GEMINI_API_KEY", "")
    parsed = parse_with_gemini(text, api_key) if api_key else None
    parser = "Gemini API" if parsed else "local parser"
    if not parsed:
        parsed = local_parse(text)
    fields = resolve_fields(parsed)
    return jsonify({"parser": parser, "fields": fields})


@app.get("/api/status")
def api_status():
    return jsonify(
        {
            "gemini_configured": bool(os.getenv("GEMINI_API_KEY", "")),
            "gemini_model": os.getenv("GEMINI_MODEL", "gemini-2.5-flash"),
            "template_exists": SOURCE_BOOK.exists(),
            "max_upload_mb": int(os.getenv("MAX_UPLOAD_MB", "12")),
            "max_text_chars": max_text_chars(),
        }
    )


@app.get("/api/suggest")
def api_suggest():
    kind = flask_request.args.get("kind", "")
    query = flask_request.args.get("q", "")[:120]
    if kind == "customer":
        candidates = ranked_matches(query, MASTER["customers"], ("name", "kana", "code"), limit=10)
    elif kind == "delivery":
        candidates = ranked_matches(query, MASTER["deliveries"], ("name", "kana", "code", "address"), limit=10)
    elif kind == "product":
        candidates = ranked_matches(query, MASTER["products"], ("name", "code"), limit=10)
    else:
        return jsonify({"error": "unknown kind"}), 400
    return jsonify(
        {
            "kind": kind,
            "query": query,
            "candidates": candidates,
            "needs_confirmation": needs_confirmation(candidates),
        }
    )


@app.post("/api/parse-image")
def api_parse_image():
    api_key = os.getenv("GEMINI_API_KEY", "")
    if not api_key:
        return jsonify({"error": "画像読み取りには環境変数 GEMINI_API_KEY が必要です。設定してアプリを再起動してください。"}), 400
    file = flask_request.files.get("image")
    if not file:
        return jsonify({"error": "画像ファイルがありません。"}), 400
    suffix = Path(file.filename or "").suffix.lower()
    mime_type = file.mimetype or mimetypes.guess_type(file.filename or "")[0] or ""
    if suffix not in ALLOWED_UPLOAD_EXTENSIONS:
        audit("upload_rejected_extension", suffix=suffix)
        return jsonify({"error": "許可されていないファイル形式です。jpg/png/webp/pdf を使ってください。"}), 400
    if not (mime_type.startswith(ALLOWED_UPLOAD_MIME_PREFIXES) or mime_type in ALLOWED_UPLOAD_MIME_TYPES):
        audit("upload_rejected_mime", mime=mime_type)
        return jsonify({"error": "許可されていないファイル形式です。"}), 400
    image_bytes = file.read()
    if not image_bytes:
        return jsonify({"error": "画像ファイルが空です。"}), 400
    if not allowed_upload_signature(image_bytes, suffix):
        audit("upload_rejected_signature", suffix=suffix, mime=mime_type)
        return jsonify({"error": "ファイルの中身が拡張子と一致しません。jpg/png/webp/pdf を使ってください。"}), 400

    audit("image_parse_requested", filename=Path(file.filename or "upload").name, size=len(image_bytes), mime=mime_type)
    parsed = parse_image_with_gemini(image_bytes, file.filename or "order.jpg", api_key)
    if not parsed:
        return jsonify({"error": "Gemini APIで画像を読み取れませんでした。画像の明るさ、ピント、APIキーを確認してください。"}), 400
    fields = resolve_fields(parsed)
    return jsonify({"parser": "Gemini Vision", "fields": fields})


@app.post("/api/save")
def api_save():
    payload = json_payload()
    if payload is None:
        return jsonify({"error": "JSON形式が正しくありません。"}), 400
    fields = payload.get("fields") or {}
    if not isinstance(fields, dict):
        return jsonify({"error": "保存データの形式が正しくありません。"}), 400
    path = save_excel(fields)
    token = create_download_token(path)
    audit("excel_saved", file=path.name)
    return jsonify({"path": str(path), "download_url": f"/download/{token}"})


@app.post("/api/quote/parse")
def api_quote_parse():
    payload = json_payload()
    if payload is None:
        return jsonify({"error": "JSON形式が正しくありません。"}), 400
    text = str(payload.get("text", "")).strip()
    if not text:
        return jsonify({"error": "見積依頼テキストが空です。"}), 400
    if len(text) > max_text_chars():
        audit("quote_parse_text_too_long", length=len(text), limit=max_text_chars())
        return jsonify({"error": f"入力が長すぎます。{max_text_chars()}文字以内にしてください。"}), 400

    api_key = os.getenv("GEMINI_API_KEY", "")
    parsed = parse_quote_with_gemini(text, api_key) if api_key else None
    parser = "Gemini API" if parsed else "local parser"
    if not parsed:
        parsed = quote_local_parse(text)
    fields = resolve_quote_fields(parsed)
    return jsonify({"parser": parser, "fields": fields})


@app.post("/api/quote/parse-image")
def api_quote_parse_image():
    api_key = os.getenv("GEMINI_API_KEY", "")
    if not api_key:
        return jsonify({"error": "画像読み取りには環境変数 GEMINI_API_KEY が必要です。設定してアプリを再起動してください。"}), 400
    file = flask_request.files.get("image")
    if not file:
        return jsonify({"error": "画像ファイルがありません。"}), 400
    suffix = Path(file.filename or "").suffix.lower()
    mime_type = file.mimetype or mimetypes.guess_type(file.filename or "")[0] or ""
    if suffix not in ALLOWED_UPLOAD_EXTENSIONS:
        audit("quote_upload_rejected_extension", suffix=suffix)
        return jsonify({"error": "許可されていないファイル形式です。jpg/png/webp/pdf を使ってください。"}), 400
    if not (mime_type.startswith(ALLOWED_UPLOAD_MIME_PREFIXES) or mime_type in ALLOWED_UPLOAD_MIME_TYPES):
        audit("quote_upload_rejected_mime", mime=mime_type)
        return jsonify({"error": "許可されていないファイル形式です。"}), 400
    image_bytes = file.read()
    if not image_bytes:
        return jsonify({"error": "画像ファイルが空です。"}), 400
    if not allowed_upload_signature(image_bytes, suffix):
        audit("quote_upload_rejected_signature", suffix=suffix, mime=mime_type)
        return jsonify({"error": "ファイルの中身が拡張子と一致しません。jpg/png/webp/pdf を使ってください。"}), 400

    audit("quote_image_parse_requested", filename=Path(file.filename or "upload").name, size=len(image_bytes), mime=mime_type)
    parsed = parse_quote_image_with_gemini(image_bytes, file.filename or "quote.jpg", api_key)
    if not parsed:
        return jsonify({"error": "Gemini APIで画像を読み取れませんでした。画像の明るさ、ピント、APIキーを確認してください。"}), 400
    fields = resolve_quote_fields(parsed)
    return jsonify({"parser": "Gemini Vision", "fields": fields})


@app.post("/api/quote/save-estimate")
def api_quote_save_estimate():
    payload = json_payload()
    if payload is None:
        return jsonify({"error": "JSON形式が正しくありません。"}), 400
    fields = payload.get("fields") or {}
    if not isinstance(fields, dict):
        return jsonify({"error": "保存データの形式が正しくありません。"}), 400
    path = save_quote_estimate_excel(fields)
    token = create_download_token(path)
    audit("quote_estimate_saved", file=path.name)
    return jsonify({"path": str(path), "download_url": f"/download/{token}"})


@app.post("/api/quote/save-request")
def api_quote_save_request():
    payload = json_payload()
    if payload is None:
        return jsonify({"error": "JSON形式が正しくありません。"}), 400
    fields = payload.get("fields") or {}
    if not isinstance(fields, dict):
        return jsonify({"error": "保存データの形式が正しくありません。"}), 400
    path = save_quote_request_excel(fields)
    token = create_download_token(path)
    audit("quote_request_saved", file=path.name)
    return jsonify({"path": str(path), "download_url": f"/download/{token}"})


@app.get("/download/<token>")
def download(token: str):
    if not re.fullmatch(r"[A-Za-z0-9_-]{20,120}", token):
        audit("download_bad_token")
        return jsonify({"error": "file not found"}), 404
    token_data = DOWNLOAD_TOKENS.get(token)
    if not token_data:
        audit("download_missing_token")
        return jsonify({"error": "file not found"}), 404
    path, expires_at = token_data
    if time.time() > expires_at or not path.exists() or path.parent != OUTPUT_DIR:
        DOWNLOAD_TOKENS.pop(token, None)
        audit("download_expired")
        return jsonify({"error": "file expired"}), 404
    audit("download", file=path.name)
    return send_file(path, as_attachment=True)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8765, debug=False)
